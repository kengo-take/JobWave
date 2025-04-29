import streamlit as st
import openpyxl as px
import const
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader

with open('config.yaml') as file:
    config = yaml.load(file, Loader=SafeLoader)

st.set_page_config(page_icon="📝")
st.markdown(const.HIDE_ST_STYLE, unsafe_allow_html=True)


authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days']
)

if st.session_state["authentication_status"]:
    ## ログイン成功
    with st.sidebar:
        st.markdown(f'## ようこそ！ *{st.session_state["name"]}さん*')
        authenticator.logout('Logout', 'sidebar')
        st.divider()

    st.header('変換（Excelのみ）')

    # ファイルアップロード
    st.error('※Excelファイルのみアップロードできます。')
    uploaded_files = st.file_uploader('変換する原稿と変更指示をアップロードしてください。', type=["xlsx", "xls"])

    # ファイルがアップロードされた場合
    if uploaded_files is not None:
        try:
            wb = px.load_workbook(uploaded_files)
            df_manuscript = wb['原稿']
            df_conversion = wb['変換指示']
            m_max = df_manuscript.max_row
            c_max = df_conversion.max_row
            if m_max != c_max:
                st.error("原稿と変換指示の数が異なります。")
            else:
                st.success("ファイルが正常に読み込まれました。")
        except Exception as e:
            st.error(f"エラー: {e}")
        
    
        # 簡単なデータ処理
        r = 1
        message_placeholder = st.empty()
        for row in df_manuscript.iter_rows(min_row=2,min_col=1):
            r += 1
            for cell in row:
                c = 0
                for col in df_conversion.columns: 
                    c += 1
                    if isinstance(cell.value, str):
                        df_conversion.cell(r, c).value = str(df_conversion.cell(r, c).value)
                        df_conversion.cell(1, c).value = str(df_conversion.cell(1, c).value)
                        cell.value = cell.value.replace(df_conversion.cell(1, c).value, df_conversion.cell(r, c).value)
            message_placeholder.write(f'*{r-1}本目/{m_max-1}の原稿を処理中です*')
        wb.save(uploaded_files)

                    
    
        # 処理結果をCSVとしてダウンロード
        st.download_button(
            label="変換した原稿をダウンロードする",
            data=uploaded_files,
            file_name='processed_data.xlsx',
        )
else:
    st.warning("**ログイン画面からログインしてください**")